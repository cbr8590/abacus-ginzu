#!/usr/bin/env python3
"""
Ginzu Model Autocomplete - Fills Damodaran high-growth valuation Excel template
using local Llama (Ollama) or Google Gemini. Upload a diligence doc or enter a company name.
Ollama is free, no API keys, no rate limits. Install: https://ollama.com
"""

import os
import re
import sys
from pathlib import Path

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

from document_ingestion import extract_text
from document_extraction import (
    extract_and_derive_from_document,
    print_validation_summary,
    print_validation_table,
)
from ginzu_config import GINZU_INPUT_MAPPING, get_cell_location
from ginzu_utils import normalize_ginzu_response

try:
    import ginzu_debug
except ImportError:
    ginzu_debug = None


def _debug(msg: str, level: str = "info"):
    if ginzu_debug:
        ginzu_debug.log(msg, level)


def _get_llm_client():
    """Ollama (local) is preferred when available. Gemini fallback when Ollama not running. Set PREFER_OLLAMA=1 for Ollama-only."""
    prefer_ollama = os.environ.get("PREFER_OLLAMA", "0").lower() in ("1", "true", "yes")

    try:
        from ollama_client import OllamaClient, _check_ollama_available
        if _check_ollama_available():
            _debug("Using Ollama (local Llama) - preferred, no API key, no limits")
            return OllamaClient(), "Ollama"
    except ImportError:
        _debug("ollama package not installed - pip install ollama")
    except Exception as e:
        _debug(f"Ollama not available: {e}")

    if prefer_ollama:
        _debug("PREFER_OLLAMA=1: Ollama not available, not falling back to Gemini", "ERROR")
        return None, None

    # Fallback to Gemini only when PREFER_OLLAMA is off
    gemini_key = os.environ.get("GEMINI_API_KEY")
    if not gemini_key:
        return None, None
    from gemini_client import GeminiClient
    _debug("Using Gemini (cloud API) - fallback")
    return GeminiClient(api_key=gemini_key), "Gemini"


def write_to_excel(file_path: str, values: dict, output_path: str = None) -> Path:
    """Write autocomplete values to the Excel file. Always outputs .xlsx. Returns the output path."""
    file_path = Path(file_path)
    out = Path(output_path) if output_path else file_path.parent / f"{file_path.stem}_autocomplete.xlsx"
    # Ensure .xlsx extension
    if out.suffix.lower() != ".xlsx":
        out = out.with_suffix(".xlsx")

    if file_path.suffix.lower() == ".xlsx":
        _write_xlsx(file_path, out, values)
    else:
        _write_xls_to_xlsx(file_path, out, values)

    return out


def run_autocomplete(
    *,
    document_path: str = None,
    company_name: str = None,
    template_path: str,
    output_path: str = None,
    instructions_path: str = None,
    dry_run: bool = False,
) -> tuple:
    """
    Run autocomplete. Returns (success, output_file_path, error_message).
    """
    try:
        from dotenv import load_dotenv
        load_dotenv()
    except ImportError:
        pass

    _debug("run_autocomplete() called")
    use_document = bool(document_path)
    if use_document:
        _debug(f"Document mode: path={document_path}")
        if not Path(document_path).exists():
            _debug(f"ERROR: Document not found: {document_path}", "ERROR")
            return False, None, f"Document not found: {document_path}"
        company = company_name or "the company in this document"
    else:
        _debug(f"Company mode: name={company_name}")
        if not company_name or not company_name.strip():
            _debug("ERROR: Company name required", "ERROR")
            return False, None, "Company name required"
        company = company_name.strip()

    tpl = Path(template_path)
    if not tpl.exists():
        _debug(f"ERROR: Template not found: {template_path}", "ERROR")
        return False, None, f"Template not found: {template_path}"
    _debug(f"Template: {template_path}")

    # Document mode: document is source of truth for ANY company
    if use_document:
        client, backend = _get_llm_client()
        if not client:
            return False, None, (
                "No LLM available.\n\n"
                "Option 1 (recommended): Install Ollama for free local Llama:\n"
                "  • Download: https://ollama.com\n"
                "  • Run: ollama pull llama3.2\n"
                "  • No API key needed, no rate limits\n\n"
                "Option 2: Set GEMINI_API_KEY in .env\n"
                "  • Get key: https://aistudio.google.com/apikey"
            )
        custom_instructions = ""
        if instructions_path and Path(instructions_path).exists():
            custom_instructions = Path(instructions_path).read_text()
        try:
            _debug("Extracting text from document...")
            doc_text = extract_text(document_path)
            _debug(f"Extracted {len(doc_text):,} chars from document")
            _debug(f"Document-first: extracting historical financials via {backend}...")
            ginzu, source = extract_and_derive_from_document(
                document_text=doc_text,
                company_name=company,
                llm_client=client,
                custom_instructions=custom_instructions,
            )
            if not ginzu:
                return False, None, "No financials extracted from document"
            values = normalize_ginzu_response(ginzu)
            print_validation_summary(source)
            print_validation_table(source, ginzu)
            _debug(f"Document-first: {len(values)} fields from source")
        except Exception as e:
            _debug(f"EXCEPTION: {type(e).__name__}: {e}", "ERROR")
            err = str(e)
            if backend == "Gemini":
                if "429" in err or "quota" in err.lower() or "ResourceExhausted" in err or "limit" in err.lower():
                    return False, None, (
                        f"Gemini 429 quota exceeded.\n\n"
                        f"Try Ollama instead (free, no limits): https://ollama.com\n"
                        f"Run: ollama pull llama3.2"
                    )
                if "404" in err or "not found" in err.lower():
                    return False, None, (
                        f"Gemini 404 Model Not Found.\n\n"
                        f"Try Ollama instead (free): https://ollama.com"
                    )
            return False, None, str(e)
    else:
        # Company mode (no document): LLM infers from company name
        client, backend = _get_llm_client()
        if not client:
            return False, None, (
                "No LLM available.\n\n"
                "Option 1 (recommended): Install Ollama for free local Llama:\n"
                "  • Download: https://ollama.com\n"
                "  • Run: ollama pull llama3.2\n"
                "  • No API key needed, no rate limits\n\n"
                "Option 2: Set GEMINI_API_KEY in .env\n"
                "  • Get key: https://aistudio.google.com/apikey"
            )
        custom_instructions = ""
        if instructions_path and Path(instructions_path).exists():
            custom_instructions = Path(instructions_path).read_text()
        try:
            _debug(f"Calling {backend} get_ginzu_inputs for company: {company}")
            values = client.get_ginzu_inputs(
                company=company,
                custom_instructions=custom_instructions,
            )
            _debug(f"{backend} returned {len(values)} fields")
        except Exception as e:
            _debug(f"EXCEPTION: {type(e).__name__}: {e}", "ERROR")
            err = str(e)
            if backend == "Gemini":
                if "429" in err or "quota" in err.lower() or "ResourceExhausted" in err or "limit" in err.lower():
                    return False, None, (
                        f"Gemini 429 quota exceeded.\n\n"
                        f"Try Ollama instead (free, no limits): https://ollama.com\n"
                        f"Run: ollama pull llama3.2"
                    )
                if "404" in err or "not found" in err.lower():
                    return False, None, (
                        f"Gemini 404 Model Not Found.\n\n"
                        f"Try Ollama instead (free): https://ollama.com"
                    )
            return False, None, str(e)

    if not values:
        _debug("ERROR: No values returned", "ERROR")
        return False, None, "No values returned from LLM"

    # Prefer company name for filename when available (e.g. Atlas_Robotics_ginzu.xls)
    def _output_base() -> str:
        generic = ("the company in this document", "the company", "")
        if company and company.strip().lower() not in (g.lower() for g in generic):
            return re.sub(r"[^\w\-]", "_", company.strip())[:50]
        if use_document and document_path:
            return Path(document_path).stem[:50]
        return "ginzu_output"

    if dry_run:
        _debug("[Dry run] Skipping Excel write")
        downloads = Path.home() / "Downloads"
        would_be = downloads / f"{_output_base()}_ginzu.xlsx"
        return True, str(would_be), None

    # Default: save to Downloads as .xlsx (e.g. Atlas_Robotics_ginzu.xlsx)
    if not output_path:
        downloads = Path.home() / "Downloads"
        output_path = str(downloads / f"{_output_base()}_ginzu.xlsx")

    _debug(f"Writing {len(values)} values to Excel...")
    out = write_to_excel(str(tpl), values, output_path)
    _debug(f"Saved to: {out}")
    return True, str(out), None


def _write_xls_to_xlsx(input_path: Path, output_path: Path, values: dict) -> None:
    """Read .xls with xlrd, write .xlsx with openpyxl (converts format)."""
    import xlrd
    from openpyxl import Workbook

    rb = xlrd.open_workbook(str(input_path))
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for sheet_idx in range(rb.nsheets):
        rs = rb.sheet_by_index(sheet_idx)
        ws = wb.create_sheet(title=rs.name[:31], index=sheet_idx)  # Excel sheet names max 31 chars
        for row in range(rs.nrows):
            for col in range(rs.ncols):
                cell = rs.cell(row, col)
                val = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    from datetime import datetime
                    from xlrd import xldate_as_tuple
                    try:
                        val = datetime(*xldate_as_tuple(val, rb.datemode))
                    except Exception:
                        pass
                ws.cell(row=row + 1, column=col + 1, value=val)

    # Apply our values
    for field_name, value in values.items():
        if value is None:
            continue
        loc = get_cell_location(field_name)
        if not loc:
            continue
        sheet_name, row, col = loc
        try:
            ws = next(ws for ws in wb.worksheets if sheet_name.lower() in ws.title.lower())
        except StopIteration:
            continue
        cell = ws.cell(row=row + 1, column=col + 1)
        cell.value = "Yes" if value is True else ("No" if value is False else value)
    wb.save(str(output_path))


def _write_xls(input_path: Path, output_path: Path, values: dict) -> None:
    import xlrd
    import xlwt
    from xlutils.copy import copy

    rb = xlrd.open_workbook(str(input_path))
    wb = copy(rb)

    for field_name, value in values.items():
        if value is None:
            continue
        loc = get_cell_location(field_name)
        if not loc:
            continue
        sheet_name, row, col = loc
        try:
            sheet_idx = next(i for i, s in enumerate(rb.sheet_names()) if sheet_name.lower() in s.lower())
        except StopIteration:
            continue
        ws = wb.get_sheet(sheet_idx)
        if isinstance(value, bool):
            ws.write(row, col, "Yes" if value else "No")
        elif isinstance(value, (int, float)):
            ws.write(row, col, value)
        else:
            ws.write(row, col, str(value))

    wb.save(str(output_path))


def _write_xls_to_xlsx(input_path: Path, output_path: Path, values: dict) -> None:
    """Read .xls with xlrd, write .xlsx with openpyxl (converts format)."""
    import xlrd
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    rb = xlrd.open_workbook(str(input_path))
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for sheet_idx in range(rb.nsheets):
        rs = rb.sheet_by_index(sheet_idx)
        ws = wb.create_sheet(title=rs.name[:31], index=sheet_idx)  # Excel sheet name max 31 chars
        for row in range(rs.nrows):
            for col in range(rs.ncols):
                cell = rs.cell(row, col)
                val = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    from datetime import datetime
                    from xlrd import xldate_as_tuple
                    try:
                        val = datetime(*xldate_as_tuple(val, rb.datemode))
                    except Exception:
                        pass
                ws.cell(row=row + 1, column=col + 1, value=val)

    # Apply our value overrides
    for field_name, value in values.items():
        if value is None:
            continue
        loc = get_cell_location(field_name)
        if not loc:
            continue
        sheet_name, row, col = loc
        try:
            ws = next(s for s in wb.worksheets if sheet_name.lower() in s.title.lower())
            cell = ws.cell(row=row + 1, column=col + 1)
            cell.value = "Yes" if value is True else ("No" if value is False else value)
        except StopIteration:
            continue

    wb.save(str(output_path))


def _write_xls_to_xlsx(input_path: Path, output_path: Path, values: dict) -> None:
    """Read .xls with xlrd, write .xlsx with openpyxl (converts format)."""
    import xlrd
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    rb = xlrd.open_workbook(str(input_path))
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for sheet_idx in range(rb.nsheets):
        rs = rb.sheet_by_index(sheet_idx)
        ws = wb.create_sheet(title=rs.name[:31], index=sheet_idx)
        for row in range(rs.nrows):
            for col in range(rs.ncols):
                cell = rs.cell(row, col)
                val = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    from datetime import datetime
                    from xlrd import xldate_as_datetime
                    try:
                        val = xldate_as_datetime(cell.value, rb.datemode)
                    except Exception:
                        pass
                ws.cell(row=row + 1, column=col + 1, value=val)

    for field_name, value in values.items():
        if value is None:
            continue
        loc = get_cell_location(field_name)
        if not loc:
            continue
        sheet_name, row, col = loc
        try:
            ws = next(s for s in wb.worksheets if sheet_name.lower() in s.title.lower())
        except StopIteration:
            continue
        cell = ws.cell(row=row + 1, column=col + 1)
        cell.value = "Yes" if isinstance(value, bool) else value

    wb.save(str(output_path))


def _write_xls_to_xlsx(input_path: Path, output_path: Path, values: dict) -> None:
    """Read .xls with xlrd, write .xlsx with openpyxl (converts format)."""
    import xlrd
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    rb = xlrd.open_workbook(str(input_path))
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for sheet_idx, sheet_name in enumerate(rb.sheet_names()):
        rs = rb.sheet_by_index(sheet_idx)
        ws = wb.create_sheet(title=sheet_name[:31])
        for row in range(rs.nrows):
            for col in range(rs.ncols):
                cell = rs.cell(row, col)
                val = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    from datetime import datetime
                    from xlrd import xldate_as_datetime
                    val = xldate_as_datetime(val, rb.datemode)
                c = ws.cell(row=row + 1, column=col + 1, value=val)
                if val is not None and val != "":
                    c.value = val

    # Apply our value overrides
    for field_name, value in values.items():
        if value is None:
            continue
        loc = get_cell_location(field_name)
        if not loc:
            continue
        sheet_name, row, col = loc
        try:
            ws = next(s for s in wb.worksheets if sheet_name.lower() in s.title.lower())
        except StopIteration:
            continue
        cell = ws.cell(row=row + 1, column=col + 1)
        cell.value = "Yes" if value is True else ("No" if value is False else value)
    wb.save(str(output_path))


def _write_xls_to_xlsx(input_path: Path, output_path: Path, values: dict) -> None:
    """Read .xls with xlrd, write .xlsx with openpyxl. Copies cell values and applies our overrides."""
    import xlrd
    from openpyxl import Workbook

    rb = xlrd.open_workbook(str(input_path))
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for sheet_idx, sheet_name in enumerate(rb.sheet_names()):
        xs = rb.sheet_by_index(sheet_idx)
        ws = wb.create_sheet(title=sheet_name[:31])
        for row in range(xs.nrows):
            for col in range(xs.ncols):
                try:
                    cell = xs.cell(row, col)
                    val = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        from datetime import datetime
                        from xlrd import xldate_as_tuple
                        t = xldate_as_tuple(val, rb.datemode)
                        val = datetime(*t)
                    if val is not None and val != "":
                        ws.cell(row=row + 1, column=col + 1, value=val)
                except Exception:
                    pass

    for field_name, value in values.items():
        if value is None:
            continue
        loc = get_cell_location(field_name)
        if not loc:
            continue
        sheet_name, row, col = loc
        try:
            ws = next(s for s in wb.worksheets if sheet_name.lower() in s.title.lower())
            cell = ws.cell(row=row + 1, column=col + 1)
            cell.value = "Yes" if value is True else ("No" if value is False else value)
        except StopIteration:
            continue

    wb.save(str(output_path))


def _write_xlsx(input_path: Path, output_path: Path, values: dict) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(str(input_path), data_only=False)
    for field_name, value in values.items():
        if value is None:
            continue
        loc = get_cell_location(field_name)
        if not loc:
            continue
        sheet_name, row, col = loc
        try:
            ws = next(ws for name, ws in [(s.title, s) for s in wb.worksheets] if sheet_name.lower() in name.lower())
        except StopIteration:
            continue
        cell = ws.cell(row=row + 1, column=col + 1)
        cell.value = "Yes" if isinstance(value, bool) else value
    wb.save(str(output_path))


def main():
    parser = __import__("argparse").ArgumentParser(
        description="Autocomplete Ginzu Excel model using Ollama (local Llama) or Gemini"
    )
    parser.add_argument("company", nargs="?", help="Company name or ticker (e.g., Tesla)")
    parser.add_argument("-t", "--template", default=os.path.expanduser("~/Downloads/higrowth.xls"), help="Excel template path")
    parser.add_argument("-o", "--output", help="Output file path")
    parser.add_argument("-i", "--instructions", help="Custom instructions file")
    parser.add_argument("-d", "--document", help="Company diligence document (PDF, Word, txt)")
    parser.add_argument("--list-fields", action="store_true", help="List input fields")
    parser.add_argument("--dry-run", action="store_true", help="Fetch but don't write Excel")
    args = parser.parse_args()

    if args.list_fields:
        for name in GINZU_INPUT_MAPPING.keys():
            loc = get_cell_location(name)
            sheet = loc[0] if loc else "?"
            print(f"  {name} ({sheet})")
        return 0

    success, out_path, err = run_autocomplete(
        document_path=args.document,
        company_name=args.company,
        template_path=args.template,
        output_path=args.output,
        instructions_path=args.instructions,
        dry_run=args.dry_run,
    )

    if not success:
        print(f"\nError: {err}")
        return 1

    if args.dry_run:
        print("\n[Dry run] Fetched values, no Excel written")
        return 0

    print(f"\nSaved to: {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
